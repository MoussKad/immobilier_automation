"""
Module d'export (CSV, Excel, SQLite) et d'alertes (email).
"""

import os
import glob
import logging
import smtplib
import sqlite3
from datetime import datetime, timedelta
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from typing import List, Optional
from pathlib import Path

import pandas as pd

logger = logging.getLogger(__name__)


class DataExporter:
    """Exporte les données vers différents formats."""

    def __init__(self, config: dict):
        self.config = config
        self.export_config = config.get("export", {})
        os.makedirs("data/results", exist_ok=True)
        os.makedirs("data/cleaned", exist_ok=True)

    def cleanup(self) -> dict:
        """
        Supprime les fichiers exportés plus vieux que export.retention_days jours.

        Fichiers concernés :
          - data/results/annonces_*.csv
          - data/results/annonces_*.xlsx
          - data/results/last_run.json  (si plus vieux que retention_days)
          - data/cleaned/annonces_clean_*.csv
          - data/raw/annonces_raw_*.json

        Retourne un dict avec le bilan : {'supprimés': [...], 'conservés': int, 'erreurs': [...]}
        """
        retention_days = int(self.export_config.get("retention_days", 7))

        if retention_days == 0:
            logger.info("🗂  Nettoyage désactivé (retention_days = 0)")
            return {"supprimés": [], "conservés": 0, "erreurs": []}

        cutoff = datetime.now() - timedelta(days=retention_days)
        logger.info(
            f"🗂  Nettoyage des fichiers antérieurs au "
            f"{cutoff.strftime('%d/%m/%Y %H:%M')} (retention_days={retention_days})"
        )

        patterns = [
            "data/results/annonces_*.csv",
            "data/results/annonces_*.xlsx",
            "data/cleaned/annonces_clean_*.csv",
            "data/raw/annonces_raw_*.json",
        ]

        supprimés = []
        erreurs = []
        conservés = 0

        for pattern in patterns:
            for filepath in glob.glob(pattern):
                try:
                    mtime = datetime.fromtimestamp(os.path.getmtime(filepath))
                    if mtime < cutoff:
                        size_kb = os.path.getsize(filepath) / 1024
                        os.remove(filepath)
                        supprimés.append(filepath)
                        logger.info(
                            f"  ✗ Supprimé : {filepath} "
                            f"({size_kb:.1f} Ko, modifié le {mtime.strftime('%d/%m/%Y')})"
                        )
                    else:
                        conservés += 1
                except OSError as e:
                    erreurs.append({"fichier": filepath, "erreur": str(e)})
                    logger.warning(f"  ⚠ Impossible de supprimer {filepath} : {e}")

        # Résumé
        if supprimés:
            logger.info(
                f"✓ Nettoyage terminé : {len(supprimés)} fichier(s) supprimé(s), "
                f"{conservés} conservé(s)"
            )
        else:
            logger.info(
                f"✓ Nettoyage terminé : aucun fichier à supprimer "
                f"({conservés} fichier(s) conservé(s))"
            )

        if erreurs:
            logger.warning(f"⚠ {len(erreurs)} erreur(s) lors du nettoyage : {erreurs}")

        return {"supprimés": supprimés, "conservés": conservés, "erreurs": erreurs}

    def _get_timestamp(self) -> str:
        return datetime.now().strftime("%Y%m%d_%H%M%S")

    def export_csv(self, df: pd.DataFrame, filename: str = None) -> str:
        """Exporte vers CSV."""
        ts = self._get_timestamp()
        filename = filename or f"data/results/annonces_{ts}.csv"
        df.to_csv(filename, index=False, encoding="utf-8-sig")
        logger.info(f"✓ CSV exporté → {filename} ({len(df)} lignes)")
        return filename

    def export_excel(self, df: pd.DataFrame,
                     df_recommandees: pd.DataFrame = None,
                     filename: str = None) -> str:
        """Exporte vers Excel avec plusieurs onglets."""
        ts = self._get_timestamp()
        filename = filename or f"data/results/annonces_{ts}.xlsx"

        with pd.ExcelWriter(filename, engine="openpyxl") as writer:
            # Onglet principal : toutes les annonces
            df.to_excel(writer, sheet_name="Toutes les annonces", index=False)

            # Onglet recommandées
            if df_recommandees is not None and len(df_recommandees) > 0:
                df_recommandees.to_excel(writer, sheet_name="Recommandées ⭐", index=False)

            # Onglet statistiques
            stats_df = self._build_stats_sheet(df)
            stats_df.to_excel(writer, sheet_name="Statistiques", index=False)

            # Mise en forme
            self._format_excel(writer, df, "Toutes les annonces")
            if df_recommandees is not None and len(df_recommandees) > 0:
                self._format_excel(writer, df_recommandees, "Recommandées ⭐")

        logger.info(f"✓ Excel exporté → {filename}")
        return filename

    def _format_excel(self, writer, df: pd.DataFrame, sheet_name: str):
        """Applique une mise en forme basique au fichier Excel."""
        try:
            from openpyxl.styles import PatternFill, Font, Alignment
            ws = writer.sheets[sheet_name]

            # En-tête en gras et coloré
            header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            # Largeur des colonnes
            col_widths = {
                "titre": 45, "prix": 15, "surface": 12, "etage": 12,
                "orientation": 15, "etat": 15, "score_global": 15,
                "categorie": 20, "lien": 50, "description": 60,
            }
            for col in df.columns:
                col_idx = df.columns.get_loc(col) + 1
                width = col_widths.get(col, 18)
                ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

            # Colorer les lignes selon le score
            if "score_global" in df.columns:
                score_col_idx = list(df.columns).index("score_global") + 1
                for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
                    try:
                        score_cell = ws.cell(row=row_idx, column=score_col_idx)
                        score = score_cell.value
                        if score is not None:
                            if score >= 80:
                                fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                            elif score >= 65:
                                fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                            else:
                                fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                            for cell in row:
                                cell.fill = fill
                    except Exception:
                        pass

        except ImportError:
            logger.warning("openpyxl styles non disponibles - export sans mise en forme")
        except Exception as e:
            logger.warning(f"Erreur mise en forme Excel : {e}")

    def _build_stats_sheet(self, df: pd.DataFrame) -> pd.DataFrame:
        """Construit un DataFrame de statistiques."""
        stats = []

        if "prix" in df.columns:
            stats.append({"Indicateur": "Prix moyen", "Valeur": f"{df['prix'].mean():,.0f} €"})
            stats.append({"Indicateur": "Prix médian", "Valeur": f"{df['prix'].median():,.0f} €"})
            stats.append({"Indicateur": "Prix minimum", "Valeur": f"{df['prix'].min():,.0f} €"})
            stats.append({"Indicateur": "Prix maximum", "Valeur": f"{df['prix'].max():,.0f} €"})

        if "surface" in df.columns:
            stats.append({"Indicateur": "Surface moyenne", "Valeur": f"{df['surface'].mean():.0f} m²"})

        if "prix_m2" in df.columns:
            stats.append({"Indicateur": "Prix/m² moyen", "Valeur": f"{df['prix_m2'].mean():,.0f} €/m²"})

        if "score_global" in df.columns:
            stats.append({"Indicateur": "Score moyen", "Valeur": f"{df['score_global'].mean():.1f}/100"})

        if "categorie" in df.columns:
            for cat, count in df["categorie"].value_counts().items():
                stats.append({"Indicateur": f"Catégorie: {cat}", "Valeur": str(count)})

        if "orientation" in df.columns:
            for ori, count in df["orientation"].value_counts().head(5).items():
                stats.append({"Indicateur": f"Orientation: {ori}", "Valeur": str(count)})

        stats.append({"Indicateur": "Total annonces", "Valeur": str(len(df))})

        return pd.DataFrame(stats)

    def export_sqlite(self, df: pd.DataFrame, db_path: str = None) -> str:
        """Exporte vers SQLite."""
        import sqlite3
        db_path = db_path or self.export_config.get("db_path", "data/results/immobilier.db")
        os.makedirs(os.path.dirname(db_path), exist_ok=True)

        conn = sqlite3.connect(db_path)
        try:
            # Ajouter timestamp de collecte si absent
            if "date_collecte" not in df.columns:
                df["date_collecte"] = datetime.now().isoformat()

            df.to_sql("annonces", conn, if_exists="append", index=False)

            # Supprimer les doublons en DB par ID
            conn.execute("""
                DELETE FROM annonces
                WHERE rowid NOT IN (
                    SELECT MAX(rowid) FROM annonces GROUP BY id
                )
            """)
            conn.commit()

            count = conn.execute("SELECT COUNT(*) FROM annonces").fetchone()[0]
            logger.info(f"✓ SQLite exporté → {db_path} ({count} entrées au total)")

        finally:
            conn.close()

        return db_path

    def export_all(self, df: pd.DataFrame,
                   df_recommandees: pd.DataFrame = None) -> dict:
        """Lance tous les exports configurés."""
        paths = {}

        if self.export_config.get("csv", True):
            paths["csv"] = self.export_csv(df)

        if self.export_config.get("excel", True):
            paths["excel"] = self.export_excel(df, df_recommandees)

        if self.export_config.get("sqlite", True):
            db_path = self.export_config.get("db_path", "data/results/immobilier.db")
            paths["sqlite"] = self.export_sqlite(df, db_path)

        return paths


class AlertSender:
    """Envoie des alertes par email."""

    def __init__(self, config: dict):
        self.config = config.get("alerts", {})

    def send_email(self, df_recommandees: pd.DataFrame,
                   excel_path: str = None) -> bool:
        """Envoie un email avec les nouvelles annonces recommandées."""
        email_config = self.config.get("email", {})

        if not email_config.get("enabled", False):
            logger.info("Email désactivé dans la config")
            return False

        if df_recommandees is None or len(df_recommandees) == 0:
            logger.info("Aucune annonce recommandée à envoyer")
            return False

        try:
            msg = MIMEMultipart("alternative")
            msg["Subject"] = f"🏠 {len(df_recommandees)} nouvelles annonces immobilières - {datetime.now().strftime('%d/%m/%Y')}"
            msg["From"] = email_config["sender"]
            msg["To"] = ", ".join(email_config["recipients"])

            # Corps HTML
            html_body = self._build_email_html(df_recommandees)
            msg.attach(MIMEText(html_body, "html", "utf-8"))

            # Pièce jointe Excel
            if excel_path and os.path.exists(excel_path):
                with open(excel_path, "rb") as f:
                    part = MIMEBase("application", "octet-stream")
                    part.set_payload(f.read())
                    encoders.encode_base64(part)
                    part.add_header(
                        "Content-Disposition",
                        f"attachment; filename={os.path.basename(excel_path)}"
                    )
                    msg.attach(part)

            with smtplib.SMTP(email_config["smtp_server"], email_config["smtp_port"]) as server:
                server.ehlo()
                server.starttls()
                server.login(email_config["sender"], email_config["password"])
                server.sendmail(
                    email_config["sender"],
                    email_config["recipients"],
                    msg.as_string()
                )

            logger.info(f"✓ Email envoyé à {email_config['recipients']}")
            return True

        except Exception as e:
            logger.error(f"Erreur envoi email : {e}")
            return False

    def _build_email_html(self, df: pd.DataFrame) -> str:
        """Génère le corps HTML de l'email."""
        rows = ""
        for _, row in df.iterrows():
            score = row.get("score_global", 0)
            color = "#27ae60" if score >= 80 else "#f39c12" if score >= 65 else "#e74c3c"

            rows += f"""
            <tr>
                <td><a href="{row.get('lien', '#')}">{row.get('titre', 'N/A')[:60]}</a></td>
                <td>{row.get('prix', 'N/A'):,.0f} €</td>
                <td>{row.get('surface', 'N/A')} m²</td>
                <td>{row.get('etage', 'N/A')}</td>
                <td>{row.get('orientation', 'N/A')}</td>
                <td style="color:{color};font-weight:bold">{score}/100</td>
                <td>{row.get('categorie', 'N/A')}</td>
            </tr>
            """

        return f"""
        <html><body>
        <h2>🏠 Nouvelles annonces immobilières</h2>
        <p>Date : {datetime.now().strftime('%d/%m/%Y %H:%M')}</p>
        <p><strong>{len(df)} annonces</strong> correspondent à vos critères.</p>
        <table border="1" cellpadding="8" style="border-collapse:collapse;font-family:Arial">
            <thead style="background:#1F4E79;color:white">
                <tr><th>Titre</th><th>Prix</th><th>Surface</th><th>Étage</th>
                <th>Orientation</th><th>Score</th><th>Catégorie</th></tr>
            </thead>
            <tbody>{rows}</tbody>
        </table>
        <p style="font-size:12px;color:gray">Généré automatiquement par ImmobilierBot</p>
        </body></html>
        """
